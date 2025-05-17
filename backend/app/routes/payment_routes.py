# app/routes/payment_routes.py

from fastapi import APIRouter, UploadFile, File, Form
from openpyxl import Workbook, load_workbook
import os

router = APIRouter()

@router.post("/register")
async def register_user(
    name: str = Form(...),
    reg_number: str = Form(...),
    email: str = Form(...),
    screenshot: UploadFile = File(...)
):
    upload_dir = "uploads"
    os.makedirs(upload_dir, exist_ok=True)

    file_path = os.path.join(upload_dir, screenshot.filename)
    with open(file_path, "wb") as f:
        f.write(await screenshot.read())

    excel_path = os.path.join(upload_dir, "registrations.xlsx")

    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        ws.append(["Name", "Registration Number", "Email", "Screenshot Filename"])
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    ws.append([name, reg_number, email, screenshot.filename])
    wb.save(excel_path)

    return {"message": "Registration successful!"}
